#!/usr/bin/python2

import pandas as pd
file = 'C:\\Users\\rchethana\\Desktop\\insights_details - Copy.xlsx'
id =[]
from openpyxl import load_workbook

wb = load_workbook(file)
sheet = wb.get_sheet_by_name('Insights')
for i in range(3, 103):
    id.append( sheet.cell( row=i, column=2 ).value )
    sheet.cell( row=i, column=3 ).value = sheet.cell( row=i, column=2 ).value
print id

wb.save(file)



