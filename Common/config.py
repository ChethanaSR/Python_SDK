#!/usr/bin/python

#############################################################################################################################################
__filename__ = "config.py"
__description__ = "contains all required configurations for execution of SDK_regession"
__author__ = "Chethana S R"
__copyright__ = "Copyright 2019"
__credits__ = ["Chethana S R"]
__version__ = "1.0"
__maintainer__ = "Chethana"
__email__ = "rchethana@shravas.com"
__status__ = "Developing Stage"
#############################################################################################################################################


#########################################################################################################################################
#Import Packages
import os
import openpyxl
import time
import dicttoxml
from datetime import datetime
from xmldiff import main
#import pandas as pd
import csv
import xml.etree.ElementTree as et
from dateutil.parser import parse
###################################################################################################################################
#Global Variable declaration
global tc_result, actual_execution_time,end_time,\
    testcase_id,tc_slno, total_number_of_tc,bln_critical_error_found , bln_sucess, criticalerror,\
    test_description, exp_execution_time, tc_execution_time,tc_remark, lh_desc, end_time, start_time,my_list, first_tc_start_time, last_tc_end_time

global access_token, xml_td_path, html_report_path,excel_path, actual_xml_file,total_tc_execution_time

global  testcase_count, total_fail ,total_pass,total_abort, total_testcase_count, current_firsttime

global username, password




###################################################################################################################################
#Constant Declarations

xml_td_path = os.environ['USERPROFILE'] + "\\Documents\\SDK_Automation\\XML-TestData"
xml_ad_path = os.environ['USERPROFILE'] + "\\Documents\\SDK_Automation\\XML-ActualData"
html_report_path =os.environ['USERPROFILE'] + "\\Documents\\SDK_Automation\\Logs\ce_report.html"
excel_path =os.environ['USERPROFILE'] + "\\Documents\\SDK_Automation\\SDK_Run.xlsx"
access_token ="fIvDy7U4kN34oouxTgyrrpupOv0GdF1T8kIS3044nQuvFLjTuF2yeIJPSZSjtwBPIwHd0Av6RLCR9R4NyOewu5GEm9HA8spTLtQ1vCYzag1GyLEF9q9vRgSOrHvTbeaY"
total_testcase_count = tc_slno = total_fail = total_pass = total_abort = 0
# bln_critical_error_found = True
# bln_sucess = True
# criticalerror= "UNEXPECTED INTRUPTION"
#bln_critical_error_found = ""
#bln_sucess = ""
#criticalerror= " "
#end_time = datetime.now()
start_date_time = datetime.now().strftime( "%Y/%m/%d" + "  " + "%I:%M:%S %p")
start_time = 0
current_firsttime = datetime.now().replace( microsecond=0 )
#total_tc_execution_time  = datetime.now().replace( microsecond=0 )
final_execution_time = 0
tc_execution_time = 0
total_tc_execution_time = 0
result_row_number = 5
col_folder_name = 2
col_function_name = 3
col_tescase_id = 4
col_runoption = 5
col_input = 6
col_descrtiption = 8
col_exec_time = 9
col_test_option = 7
sheetname = "TC_Details"
result_list = False
tc_remark = ""

username = 'rchethana@shravas.com'
password = 'ZSErfv123$%^'

###################################################################################################################################


#Function to read Excel File
def excel_read():
    tc_slno = 0
    my_list =[]
    list = []
    workbook = openpyxl.load_workbook( excel_path )
    worksheet = workbook.get_sheet_by_name( sheetname )
    for row in range( 1, worksheet.max_row ):
        if worksheet.cell(row=row, column=col_runoption).value == "Yes":
            #slno = slno + 1
            start_time = datetime.now()
            folder_name = worksheet.cell( row=row, column=col_folder_name ).value
            func_name = worksheet.cell( row=row, column=col_function_name ).value
            testcase_id = worksheet.cell( row=row, column=col_tescase_id ).value
            test_input = worksheet.cell( row=row, column=col_input ).value
            test_data_option = str( worksheet.cell( row=row, column=col_test_option ).value )
            test_description = worksheet.cell( row=row, column=col_descrtiption ).value
            exp_execution_time = str(worksheet.cell( row=row, column=col_exec_time ).value)
            input_string = folder_name+"$"+ func_name+"$"+ testcase_id +"$"+ str(test_input) +"$"+ str(test_description) +"$"+ exp_execution_time+"$"+ test_data_option
            my_list.append(input_string)
            #print my_list
            #function_call(folder_name, func_name,testcase_id, test_input )
            folder_name_end = worksheet.cell( row=row, column=col_folder_name ).value
        # if worksheet.max_row == row and folder_name_end == "END" :
        #     criticalerror = "Execution Completed Successfully."
        # else:
        #     criticalerror = "Execution gracefully terminated."
    return my_list

##########################################################################################################################################################
#Get Total number of TestCase to be executed.
def get_total_number_testcase():
        count = 0
        workbook = openpyxl.load_workbook( excel_path )
        worksheet = workbook.get_sheet_by_name( sheetname )
        for row in range( 1, worksheet.max_row ):
            if worksheet.cell( row=row, column=col_runoption ).value == "Yes":
                count = count+1
        #total_testcase_count = count
        return count
############################################################################################################################################
#Function to convert Dictinory to xml File

def xml_write(folder_name, result,testcase_id):
    xml_data = dicttoxml.dicttoxml( result, attr_type=False, root=True )
    #path = xml_ad_path + "\\" + folder_name + "\\" + datetime.now().strftime( "%Y-%m-%d" + "," + "%I-%M-%S %p" )
    path = xml_ad_path + "\\" + folder_name + "\\" + datetime.now().strftime( "%Y-%m-%d"  )
    if not os.path.exists( path ):
        os.makedirs( path )
    #os.mkdir( path )
    #file_path = path + "\\" + datetime.now().strftime( "%Y-%m-%d" + "," + "%I-%M-%S %p" ) + ".xml"
    file_path = path + "\\" + testcase_id + ".xml"
    #print xml_data
    f = open(file_path, 'wb' )
    f.write( xml_data )
    f.close()
    return file_path

######################################################################################################################################################
#function to compare Actual XML and TestData XML File

def xml_comp(actual_xml_file, folder_name,test_input):

    expected_xml_file = xml_td_path +"\\" + folder_name + "\\" + test_input + ".xml"
    #print actual_xml_file
    #print expected_xml_file
    #Compare Actaul and Expected XML files
    diff = main.diff_files( expected_xml_file, actual_xml_file)
    required_node = []
    required_text = []
    for i in range( len( diff ) ):
        if str( type( diff[i] ) ) != "<class 'xmldiff.diff.MoveNode'>":
            required_node.append( diff[i].node )
            required_text.append( diff[i].text )
    #print(required_node)
    #print(required_text)

    my_path = []
    for i in range( len( required_node ) ):
        #my_path.append( required_node[i].replace( "//*/item[1]/", "" ) )
        my_path.append(required_node[i][slice( required_node[i].find( "item") + 8, None, None )])

    #print my_path

    tree = et.parse( expected_xml_file)
    root = tree.getroot()
    result = []
    for i in range( len( my_path ) ):
        for j in root.findall( ".//item/" + str( my_path[i] ) ):
            result.append( "Error occurred at the following node --> " + my_path[i] + "||The values are : '" + j.text + "' <> '" + required_text[i] + "'" )
            #result.append("Data Verification failed in XML file comparison  at the following node --> " + my_path[i] + "     The values are : '"  + j.text + "' <> '" +  required_text[i] + "'")
            #print ("Data Verification failed in XML file comparison  at the following node --> " + my_path[i] + "\n" + "values are : '"  + j.text + "' <> '" +  required_text[i] + "'"  +"\n")
            break
    return result


##########################################################################################################################################3
#Get Total Count of a testcase
total_testcase_count = get_total_number_testcase()

############################################################################################################################################

#total_tc_execution_time = tc_execution_time + current_firsttime

def get_total_tc_execution_time (tc_execution_time):
    final_execution_time =  tc_execution_time
    return final_execution_time


##################################################################################################################################################
#Function to Read CSV file and Compare with obtained XML File

# def xml_csv_comp(filename, folder_name, test_input):
#
#     #xml_file_path =
#     #print filename
#     tree = et.parse(filename )
#     root = tree.getroot()
#
#     series_id  = []
#     status = []
#     series_code = []
#     last_update_time = []
#     end_date = []
#     start_date = []
#     country = []
#     number_obs = []
#     source= []
#     frequency = []
#     unit = []
#     country_mismatch = []
#     series_id_mismatch = []
#     status_mismatch = []
#     series_code_mismatch = []
#     last_update_time_mismatch = []
#     end_time_mismatch = []
#     start_date_mismatch = []
#     number_obs_mismatch = []
#     source_mismatch = []
#     frequency_mismatch = []
#     unit_mismatch = []
#
#     xml_list = ["entity_id", "item[1]/seriesCode", "status/name", "last_update_time", "end_date", "country/name",
#                 "number_of_observations", "start_date", "source/name", "frequency/name", "unit/name"]
#
#     for i in range( len( xml_list ) ):
#         for name in root.findall( ".//" + str( xml_list[i] ) ):
#             if xml_list[i] == "entity_id": series_id.append( name.text )
#             elif xml_list[i] == "item[1]/seriesCode": series_code.append( name.text )
#             elif xml_list[i] == "status/name": status.append( name.text )
#             elif xml_list[i] == "last_update_time": last_update_time.append( name.text )
#             elif xml_list[i] == "end_date": end_date.append( name.text )
#             elif xml_list[i] == "country/name": country.append( name.text )
#             elif xml_list[i] == "number_of_observations": number_obs.append( name.text )
#             elif xml_list[i] == "start_date": start_date.append( name.text )
#             elif xml_list[i] == "source/name": source.append( name.text )
#             elif xml_list[i] == "frequency/name":  frequency.append( name.text )
#             elif xml_list[i] == "unit/name": unit.append( name.text )
#
#
#     #Read Csv File
#     csv_file = xml_td_path +"\\" + folder_name + "\\" + test_input + ".csv"
#     df = pandas.read_csv( csv_file )
#     row_count = sum( 1 for line in open( csv_file ) )
#     country_cnt = frequency_cnt = unit_cnt = source_cnt = status_cnt = series_id_cnt = series_code_cnt = number_obs_cnt = last_update_time_cnt = 0
#     start_date_cnt = end_date_cnt = 0
#     #result_list  = []
#
#
#     for i in range( len( series_id ) ):
#         for j in range( row_count - 1 ):
#             key = str( df['Series_Key'][j] )
#             if key == "Region" and df[series_id[i]][j] != country[i]:
#                 text = "For Series id: " + series_id[i] + ":  Required Country is: " + df[series_id[i]][j] + " obtained value  is:  " + country[i] + "\n"
#                 country_mismatch.append(text)
#             if key == "Frequency" and df[series_id[i]][j] != frequency[i]:
#                 text = "For Series id: " + series_id[i] + ":  Required frequency is: " + df[series_id[i]][j] + " obtained value  is:  " + frequency[i] + "\n"
#                 frequency_mismatch.append( text )
#             if key == "Unit" and df[series_id[i]][j] != unit[i]:
#                 text = "For Series id: " + series_id[i] + ":  Required unit is: " + df[series_id[i]][j] + " obtained value  is:  " + unit[i] + "\n"
#                 unit_mismatch.append( text )
#             if key == "Source" and df[series_id[i]][j] != source[i]:
#                 text = "For Series id: " + series_id[i] + ":  Required source is: " + df[series_id[i]][j] + " obtained value  is:  " + source[i] + "\n"
#                 source_mismatch.append( text )
#             if key == "Status" and df[series_id[i]][j] != status[i]:
#                 text = "For Series id: " + series_id[i] + ":  Required status is: " + df[series_id[i]][j] + " obtained value  is:  " + status[i] + "\n"
#                 status_mismatch.append( text )
#             if key == "Series ID" and series_id[i] not in df[series_id[i]][j]:
#                 text = "For Series id: " + series_id[i] +  " obtained value  is:  " + series_id[i] + "\n"
#                 series_id_mismatch.append( text )
#             if key == "SR Code" and df[series_id[i]][j] != series_code[i]:
#                 text = "For Series id: " + series_id[i] + ":  Required series_code is: " + df[series_id[i]][j] + " obtained value  is:  " + series_code[i] + "\n"
#                 series_code_mismatch.append( text )
#             if key == "No. of Obs" and df[series_id[i]][j] != number_obs[i]:
#                 text = "For Series id: " + series_id[i] + ":  Required number_obs is: " + df[series_id[i]][j] + " obtained value  is:  " + number_obs[i] + "\n"
#                 number_obs_mismatch.append( text )
#             if key == "Last Update Time":
#                 num = last_update_time[i].find( "T" )
#                 datestring = parse( last_update_time[i][:num] )
#                 datestring = datestring.strftime( '%d/%m/%Y' )
#                 if df[series_id[i]][j] != datestring:
#                     text = "For Series id: " + series_id[i]+ ":  Required Last Update Time is: "  + df[series_id[i]][j] +" obtained value  is:  " + datestring + "\n"
#                     last_update_time_mismatch.append( text )
#             if key == "First Obs. Date":
#                 datestring = parse( start_date[i] )
#                 datestring = datestring.strftime( '%b-%y' )
#                 if df[series_id[i]][j] != datestring:
#                     text = "For Series id: " + series_id[i]+ ":  Required First Obs. Date is: "  + df[series_id[i]][j] +" obtained value  is:  " + datestring + "\n"
#                     start_date_mismatch.append( text )
#             if key == "Last Obs. Date":
#                 datestring = parse( end_date[i] )
#                 datestring = datestring.strftime( '%b-%y' )
#                 if df[series_id[i]][j] != datestring:
#                     text = "For Series id: " + series_id[i]+ ":  Required Last Obs. Date is: "  + df[series_id[i]][j] +" obtained value  is:  " + datestring + "\n"
#                     end_time_mismatch.append( text )
#
#     mismatch = [country_mismatch, series_id_mismatch, status_mismatch, series_code_mismatch, last_update_time_mismatch, end_time_mismatch, start_date_mismatch, number_obs_mismatch, source_mismatch, frequency_mismatch,unit_mismatch]
#     return mismatch
#
#
#
#
#     #return bln_critical_error_found , bln_sucess

##################################################################################################################################################
#Function to Read CSV file and Compare with obtained XML File for given filter option
def xml_csv_comp_filter( filename, folder_name,test_input, filteroption ):
    tree = et.parse( filename )
    root = tree.getroot()

    # Read Csv File
    csv_file = xml_td_path + "\\" + folder_name + "\\" + test_input + ".csv"
    #if str( first ) in ("a"xml_td_path , "e", "i", "o", "u"):

    #if filteroption == "SeriesCode" or filteroption == "Mnemonic" or filteroption == "SRCode" :
    if filteroption in ("SeriesCode", "Mnemonic", "SRCode"):
        filterdata = root.findall( ".//item[1]/seriesCode" )
        verifyoption = "SR Code"
    if filteroption in ("Active", "Disconitnued", "Rebased"):
        filterdata = root.findall( ".//status/name" )
        verifyoption = "Status"
    if filteroption in ("Daily", "Weekly", "Monthly", "Quaterly", "Half-yearly", "Annual"):
        filterdata = root.findall( ".//frequency/name" )
        verifyoption = "Frequency"
    if filteroption in ("Unit"):
        filterdata = root.findall( ".//unit/name" )
        verifyoption = "Unit"
    if filteroption in ("Region"):
        filterdata = root.findall( ".//country/name" )
        verifyoption = "Region"
    if filteroption in ("Source"):
        filterdata = root.findall( ".//source/name" )
        verifyoption = "Source"









    filetrvalue =[]
    filetrvalue_csv =[]
    for num in range( len( filterdata ) ): filetrvalue.append( filterdata[num].text )

    #Read Csv File and Get SR Code
    with open( csv_file ) as csvfile:
        readCSV = csv.reader( csvfile, delimiter=',')
        for row in readCSV:
            if row[0] == verifyoption:
                for num in range (1, len(filetrvalue)+1): filetrvalue_csv.append(row[num])
                break


    intcun = 0
    failedsrcode =[]
    bln_sucess = False


    for i in range(len(filetrvalue)):
        for j in range (len(filetrvalue)):
            if filteroption in("Half-yearly", "Annual") :
                if filetrvalue_csv[j].find(filteroption) >=0 and filetrvalue[j] in("Semiannually", "Yearly"):
                    intcun += 1
                    break

            elif filteroption in ("Daily", "Weekly", "Monthly", "Quaterly"):
                if filetrvalue_csv[j].find(filetrvalue[i])>= 0 :
                    intcun+=1
                    break
            else:
                 if filetrvalue[i] == filetrvalue_csv[j]:
                     intcun+=1
                     break



    if intcun == len(filetrvalue) :return True






    # if filteroption == "SeriesCode":
    #     srcode = []
    #     srcode_csv =[]
    #     seriescode = root.findall( ".//item[1]/seriesCode" )
    #
    #
    #     #Get Series Id
    #     for num in range(len( seriescode ) ):srcode.append( seriescode[num].text )
    #     print srcode
    #
    #     #Read Csv File and Get SR Code
    #     with open( csv_file ) as csvfile:
    #         readCSV = csv.reader( csvfile, delimiter=',')
    #         for row in readCSV:
    #             if row[0] == "SR Code":
    #                 for num in range (1, len(srcode)+1): srcode_csv.append(row[num])
    #
    #     intcun = 0
    #     failedsrcode =[]
    #     bln_sucess = False
    #     for i in range(len(srcode)):
    #         for j in range (len(srcode)):
    #             if srcode[i] == srcode_csv[j]:intcun+=1
    #
    #
    #
    #     if intcun == len(srcode) :return True



def xml_csv_comp_filter_for_timepoint(filename, folder_name, test_input,tescas_id):
    tree = et.parse( filename )
    root = tree.getroot()

    blnFalg = False
    series_id_xml = []
    series_id_cvs = []
    time_point_value_xml = []
    time_point_date_xml =[]
    time_point_value_csv = []
    time_point_date_csv =[]


    #series = root.findall( ".//entity_id" )
    for name in root.findall( ".//entity_id" ): series_id_xml.append(name.text)
    # time_point_date = []
    # time_point_value = []
    #
    # i =0
    # for id in root.findall( ".//entity_id" ):
    #     if series_id_xml[i] == id.text:
    #         #print ".//item[" +str(i+1)+"/time_points/item/date"
    #         for name in root.findall (".//item[" +str(i+1)+"]/time_points/item/date"): time_point_date.append(name.text)
    #         for name in root.findall( ".//item[" + str( i+1 ) + "]/time_points/item/value" ): time_point_value.append(name.text )
    #         time_point_value_xml.append(time_point_value)
    #         time_point_date_xml.append(time_point_date)
    #         time_point_date = []
    #         time_point_value =[]
    #         i +=1
    # print (series_id_xml)
    # print (time_point_value_xml)
    # print (time_point_date_xml)



            # Read Csv File
    csv_file = xml_td_path + "\\" + folder_name + "\\" +  tescas_id  + ".csv"

    # with open( csv_file ) as csvfile:
    #     readCSV = csv.reader( csvfile, delimiter=',' )
    #     for row in readCSV:
    #         if row[0] == "Series ID":
    #             for num in range( 1, len( series ) + 1 ): series_id_cvs.append( row[num] )

    f = open( csv_file )
    csv_f = csv.reader( f )
    for row in csv_f:
        if row[0] == "Series ID":
            for col in range( 1, len( series_id_xml ) + 1 ): series_id_cvs.append( row[col] )

    # with open( csv_file ) as f:
    #     reader = csv.reader( f, delimiter=' ', quotechar='|' )
    #
    #     data = []
    #     for row in reader:  # Number of rows including the death rates
    #         for col in (2, 3,4, 5):  # The columns I want read   B and D
    #             data.append( row )
    #             #print data
    #             #*data.append( col )
    #     for item in data:
    #         print(item)
    #         #break
    #         #print data



    datecnt = 0
    blnFalg = False
    error_result =[]

    tree = et.parse( filename )
    root = tree.getroot()
    #for id in range( len( series_id_xml ) ): series_id_xml.append( series_id_xml[id].text )
    for I in range( 1, (len( series_id_xml )+1) ):
        date = root.findall( ".//item[" + str(I) + "]/time_points/item/date" )
        value = root.findall( ".//item["+str(I)+"]/time_points/item/value" )
        J= 0
        for J in range (len(date)):
            year_in_date = datetime.strptime( date[J].text, '%Y-%m-%d' ).year
            with open( csv_file ) as csvfile:
                readCSV = csv.reader( csvfile, delimiter=',' )
                for K in range(len(series_id_xml)):
                    if series_id_cvs[K] == series_id_xml[I-1]:
                        col = K+1
                        break



                for row in readCSV:
                    if row[0] == str( year_in_date ):
                        if round(float(value[J].text))== round(float(row[col])):
                            blnFalg = True
                            break
                        else:
                            error_result.append("Date Value is (" + row[col] + ')is not Matched for date ' + year_in_date+ "for series Id " +series_id_xml[I-1] + " " )
                            break

    print (error_result)
    if error_result == []: return blnFalg
    else: error_result










# def xml_csv_comp_filter_for_timepoint(filename, folder_name, test_input):
#     tree = et.parse( filename )
#     root = tree.getroot()
#     count = test_input.split(",")
#     print test_input
#
#     # Read Csv File
#     csv_file = xml_td_path + "\\" + folder_name + "\\" + test_input + ".csv"
#
#
#     verifyoption = "No. of Obs"
#     series = root.findall( ".//entity_id" )
#     series_id = []
#
#
#
#     series_date =[]
#     series_value = []
#     series_id_cvs = []
#     I = 1
#     for id in range (len(series)):
#         filterdate = []
#         filetrvalue = []
#         date = root.findall( ".//item[" +str(I)+ "]/time_points/item/date" )
#         value = root.findall( ".//item["+str(I)+"]/time_points/item/value" )
#         year_in_date = datetime.strptime( date[id].text, '%Y-%m-%d' ).year
#         series_id.append( series[id].text )
#         with open(csv_file ) as csvfile:
#             readCSV = csv.reader( csvfile, delimiter=',' )
#             for row in readCSV:
#                 if row[0] == "Series ID":
#                     for num in range( 1, len( series ) + 1 ): series_id_cvs.append( row[num] )
#
#                 if row[0]== str(year_in_date) and series_id_cvs[id]== series[id].text:
#                     print value[id].text
#
#         #I+=1
#         # for J in range (len(date)):
#         #     filterdate.append(date[J].text)
#         #     filetrvalue.append(value[J].text)
#         #     col = 2
#         #     row_number =0
#         # with open( csv_file ) as csvfile:
#         #     readCSV = csv.reader( csvfile, delimiter=',' )
#         #     for row in readCSV:
#         #         if row[col]== series[id].text:
#         #             col_num = col
#         #             col +=1
#         #         if row[0]==
#         #
#         #
#         # series_date.append(filterdate)
#         # series_value.append(filetrvalue)
#
#     # # Read Csv File and Get SR Code
#     # P= 1
#     # with open( csv_file ) as csvfile:
#     #     readCSV = csv.reader( csvfile, delimiter=',' )
#     #     for row in readCSV:
#     #         if row[0]== "Series ID":
#     #             for num in range( 1, len( series ) + 1 ):
#     #                 series_id_cvs.append(row[num])
#     #                 col-num = row[num]
#     #
#
#
#
#
#
#
#
#
#
#
#
#     #filetrvalue = [filetrvalue.append( value[I][J].text )[for I in range (len(series))]for J in range(len(date) ) ]
#     # filterdate =[[]]
#     # filetrvalue = [[]]
#     # filetrvalue_csv = []
#
#     # for I in range( len( series ) ):
#     #     for J in range( len( date ) ):
#     #         filterdate[I].append([series[I].text[date[J].text]])
#     #         filetrvalue.append( [series[I].text[value[J].text]])
#     #     series_id.append( series[I].text )
#
################################################################################################################################3
def xml_csv_comp_new(filename,folder_name,test_input,tescas_id ):


    tree = et.parse( filename )
    root = tree.getroot()

    # Read Csv File
    csv_file = xml_td_path + "\\" + folder_name + "\\" + tescas_id +".csv"
    #print csv_file
    #print filename

    region_xml = []
    frequency_xml = []
    unit_xml = []
    source_xml = []
    status_xml = []
    series_id_xml = []
    sr_code_xml = []
    first_obs_date_xml = []
    last_obs_date_xml = []
    no_of_obs_xml = []

    xml_list = ["entity_id", "layout/item/series_code", "status/name", "last_update_time", "end_date",
                "country/name", "number_of_observations", "start_date", "source/name", "frequency/name", "unit/name"]

    for i in range( len( xml_list ) ):
        for name in root.findall( ".//" + str( xml_list[i] ) ):
            if xml_list[i] == "entity_id":
                series_id_xml.append( name.text )
            elif xml_list[i] == "layout/item/series_code":
                sr_code_xml.append( name.text )
            elif xml_list[i] == "status/name":
                status_xml.append( name.text )
            elif xml_list[i] == "end_date":
                last_obs_date_xml.append( name.text )
            elif xml_list[i] == "country/name":
                region_xml.append( name.text )
            elif xml_list[i] == "number_of_observations":
                no_of_obs_xml.append( name.text )
            elif xml_list[i] == "start_date":
                first_obs_date_xml.append( name.text )
            elif xml_list[i] == "source/name":
                source_xml.append( name.text )
            elif xml_list[i] == "frequency/name":
                frequency_xml.append( name.text )
            elif xml_list[i] == "unit/name":
                unit_xml.append( name.text )

    #print(series_id_xml)

    # i = 0
    # series_id = []
    # sr_code = []
    # [series_id.append( name.text ) for name in root.findall( ".//entity_id" )]
    #
    #print ("Series Id Lenght = " , len(series_id))
    #print ("***********************************")

    # for name in root.findall( ".//entity_id" ):
    #     if series_id[i] == name.text:
    #         i += 1
    #         #print (".//item["+str(i)+"]/layout/item/series_code")
    #         for series_code_xml in root.findall( ".//item[" + str( i ) + "]/layout/item/series_code" ):
    #             sr_code.append( series_code_xml.text )
    #         #print sr_code
    # #
    #         sr_code_xml.append( sr_code )
    #         #print sr_code_xml
    #         sr_code = []
    #     #print sr_code_xml
    #cnt =  len(region_xml)+1
    #print ("SR_Code_XML Leangh", len(sr_code_xml))
    # csv file reading


    f = open( csv_file )
    csv_f = csv.reader( f )

    region_csv = []
    frequency_csv = []
    unit_csv = []
    source_csv = []
    status_csv = []
    series_id_csv = []
    sr_code_csv = []
    first_obs_date_csv = []
    last_obs_date_csv = []
    no_of_obs_csv = []

    blnregon = blnfrequency = blnunit = blnsource = blnstatus = blnseries_id = blnsr_code = False
    blnfirst_obs_date = blnlast_obs_date = blnno_of_obs = False

    region_cnt = frequency_cnt = unit_cnt = source_cnt = status_cnt = series_id_cnt = sr_code_cnt = 0
    first_obs_date_cnt = last_obs_date_cnt = no_of_obs_cnt = 0
    #print csv_f
    #time.sleep(1)
    # for row in csv_f:
    #     print row

    for row in csv_f:
        if row[0] == "Region":
            for col in range( 1, len( region_xml ) + 1  ): region_csv.append( row[col] )
            #print "Region = " , region_csv
        if row[0] == "Frequency":
            for col in range( 1, len( region_xml ) + 1 ): frequency_csv.append( row[col] )
            #print "Frefrequency_csv
        if row[0] == "Unit":
            for col in range( 1, len( region_xml ) + 1 ): unit_csv.append( row[col] )
            #print unit_csv
        if row[0] == "Source":
            for col in range( 1, len( region_xml ) + 1 ): source_csv.append( row[col] )
            #print source_csv
        if row[0] == "Status":
            for col in range( 1, len( region_xml ) + 1 ): status_csv.append( row[col] )
            #print status_csv
        if row[0] == "Series ID":
            for col in range( 1, len( region_xml ) + 1 ): series_id_csv.append( row[col] )
            #print series_id_csv
        if row[0] == "SR Code":
            for col in range( 1, len( region_xml ) + 1 ): sr_code_csv.append( row[col] )
            #print sr_code_csv
        if row[0] == "First Obs. Date":
            for col in range( 1, len( region_xml ) + 1 ): first_obs_date_csv.append( str( row[col] ) )
            #print first_obs_date_csv
        if row[0] == "Last Obs. Date":
            for col in range( 1, len( region_xml ) + 1 ): last_obs_date_csv.append( str( row[col] ) )
            break


    print(len(sr_code_csv))
    error_result = []
    for i in range( len( region_xml ) ):
        for j in range( len( region_xml ) ):
            if series_id_csv[i].find( series_id_xml[j] ) >= 0:
                #if region_csv[i] != region_xml[j]: error_result.append(
                if region_csv[i].find( region_xml[j] ) < 0 :
                    error_result.append( "Region (" + region_xml[j] + ')is mismatched for series  Id' + series_id_xml[j]+" " )
                if unit_csv[i] != unit_xml[j]:     error_result.append(
                    "Unit (" + unit_xml[j] + ')is mismatched for series Id ' + series_id_xml[j]+" " )
                if source_csv[i] != source_xml[j]: error_result.append(
                    "Source (" + source_xml[j] + ')is mismatched for series Id ' + series_id_xml[j]+" " )
                if status_csv[i] != status_xml[j]: error_result.append(
                    "Status (" + status_xml[j] + ')is mismatched for series Id ' + series_id_xml[j]+" " )
                if series_id_csv[i].find( series_id_xml[j] ) < 0: error_result.append(
                    "Series_ID (" + series_id_xml[j] + ')is mismatched for series Id ' + series_id_xml[j]+" " )
                #if sr_code_csv[i] != sr_code_xml[j]: error_result.append(
                    #"SR Code is (" + sr_code_xml[j] + ')is mismatched for series Id ' + series_id_xml[j]+" " )



                #Verifing first_obs_date
                required_date = first_obs_date_xml[j]
                required_date = required_date.split( "-" )
                obtained_date = str( required_date[1] + "/" + required_date[0] )
                obtained_date1 = str( required_date[2] + "/" +required_date[1] + "/" + required_date[0] )
                obtained_date2 = str( required_date[0] )
                if (first_obs_date_csv[i]  == obtained_date) or \
                   (first_obs_date_csv[i]  == obtained_date1) or\
                   (first_obs_date_csv[i]  == obtained_date2):first_obs_date_cnt +=1
                else:error_result.append("First Obs. Date (" + first_obs_date_xml[j] + ')is mismatched for series Id ' + series_id_xml[j] + " " )


               #Verifing last_obs_date
                required_date = last_obs_date_xml[j]
                required_date = required_date.split( "-" )
                obtained_date = str( required_date[1] + "/" + required_date[0] )
                obtained_date1 = str( required_date[2] + "/" + required_date[1] + "/" + required_date[0] )
                obtained_date2 = str( required_date[0] )
                if (last_obs_date_csv[i]  == obtained_date) or \
                   (last_obs_date_csv[i]  == obtained_date1) or\
                   (last_obs_date_csv[i]  == obtained_date2): last_obs_date_cnt +=1
                #else: error_result.append("Last Obs. Date (" + last_obs_date_xml[j] + ')is mismatched for series Id ' + series_id_xml[j] + " " )


                if (frequency_csv[i].find( "Annual" ) >= 0 and frequency_xml[j] == 'Yearly') or \
                        (frequency_csv[i].find( "Monthly" ) >= 0 and frequency_xml[j] == 'Monthly')  or \
                        (frequency_csv[i].find( "Quarterly" ) >= 0 and frequency_xml[j] == 'Quarterly')  or \
                        (frequency_csv[i].find( "Weekly" ) >= 0 and frequency_xml[j] == 'Weekly')  or \
                        (frequency_csv[i].find( "Daily" ) >= 0 and frequency_xml[j] == 'Daily') or \
                        (frequency_csv[i].find( "Half-yearly") >= 0 and frequency_xml[j] == 'Semiannually'): frequency_cnt += 1
                else: error_result.append("Frequency (" + frequency_xml[j] + ')is mismatched for series Id ' + series_id_xml[j] + " " )

    sr_code_xml_cnt = 0
    for i in range( len( sr_code_csv ) ):
        for j in range( len( sr_code_xml ) ):
            if sr_code_csv[i] == sr_code_xml[j]: sr_code_xml_cnt += 1
    print (sr_code_xml_cnt)

    if len( sr_code_csv ) != sr_code_xml_cnt:
        error_result.append("SR Code is (" + sr_code_csv[i] + ')is not exists for series id' + series_id_xml[j] + " " )


    #if frequency_cnt != len( series_id_xml ):print (error_result)

    print (error_result)
    error_result_string = ""
    if error_result == []: error_result_string = ""
    else:
        for x in error_result: error_result_string += x

    return error_result_string



    #if error_result != []: return error_result_string




