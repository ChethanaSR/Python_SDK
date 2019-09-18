# from ceic_api_client.api_client import ApiClient
# from ceic_api_client.apis.dictionary_api import DictionaryApi
# from ceic_api_client.apis.series_api import SeriesApi
# from ceic_api_client.apis.insights_api import InsightsApi
#
# access_token = "ap3PspPhdMM6omH4wxD6LfzI56y5H4D7DDjhNAhkYFHgqjv7bwEYjMR6qnZ1RM0Mc8nkpK0PodwfJCHPaVqLYxshpDW1IycAUcqb6O3QQ13sGqBCzBsYND4FHpoi5YST"
#
# api_client = ApiClient( header_name="Authorization", header_value=access_token )
# series_api = SeriesApi( api_client=api_client )
#
# key_words = ["4772092e-d7d3-48f0-a513-9dc02d569fc9"]
# series_search_result = series_api.search_series( keyword=", ".join( key_words ) )
#
# # Access series search results data
# series_search_items = series_search_result.data.items
# print series_search_items



from ceic_api_client.api_client import ApiClient
from ceic_api_client.apis.dictionary_api import DictionaryApi
from ceic_api_client.apis.series_api import SeriesApi
from ceic_api_client.apis.insights_api import InsightsApi
import dicttoxml
import xml.etree.ElementTree as et
from openpyxl import load_workbook
file = 'C:\\Users\\rchethana\\Documents\\SDK_Automation\\XML-TestData\\Metadata\\insights_details.xlsx'
file1 = open("C:\\Users\\rchethana\\Desktop\\notepad_python.txt","w")
wb = load_workbook(file)
sheet = wb.get_sheet_by_name('Insights')

access_token = "oC6OX7cRKs386i8KhqkHVi5BOgRCVMAkUVmL1EQlrKz8qNztEkUoLhior6c2RYBmhkTUenYrLBM8Kzl4R7cu2c4ywikS1F4cevkHegEKEdIEhoyYMzkgKXAWXqwzHyC8"
#example_insight_id = "f8e64f6b-003a-414c-895b-99b7e7dca317"



api_client = ApiClient(header_name="Authorization", header_value=access_token)
insights_api = InsightsApi(api_client=api_client)

for i in range(40, 102):
    example_insight_id = sheet.cell( row=i, column=2 ).value
    insights_result = insights_api.get_insight_series_metadata(id="7839a60b-57be-415b-b35c-7cfc7f3dbd44").to_dict()
#   print(insights_result)
    xml_data = dicttoxml.dicttoxml( insights_result, attr_type=False, root=True )
    file_path = "C:\\Users\\rchethana\\Documents\\SDK_Automation\\XML-ActualData\\Metadata\\2019-07-10\Chethana.xml"
    f = open( file_path, 'w' )
    f.write( xml_data )
    f.close()
    series_id_xml = []
    tree = et.parse( file_path )
    root = tree.getroot()

    for name in root.findall( ".//metadata/id" ):
        series_id_xml.append( name.text )

    series_id_xml_string =""
    if series_id_xml == []:
        series_id_xml_string = ""
    else:
        for x in series_id_xml: series_id_xml_string + x + ", "

    sheet.cell( row=i, column=3 ).value = series_id_xml_string
    print series_id_xml
    file1.writelines( series_id_xml )
    sheet.cell( row=i, column=4 ).value = len( series_id_xml )
    wb.save( file )


    #print (series_id_xml)
    #print len( series_id_xml )

wb.save(file)
file1.close()



