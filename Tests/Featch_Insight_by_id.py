from ceic_api_client.api_client import ApiClient
from ceic_api_client.apis.dictionary_api import DictionaryApi
from ceic_api_client.apis.series_api import SeriesApi
from ceic_api_client.apis.insights_api import InsightsApi

import dicttoxml
import xml.etree.ElementTree as et
from openpyxl import load_workbook
file = 'C:\\Users\\rchethana\\Documents\\SDK_Automation\\XML-TestData\\Metadata\\insights_by_id.xlsx'
wb = load_workbook(file)
sheet = wb.get_sheet_by_name('Insight_by_ID')

# access_token = "yUmpy171bzDrqPyE6veMScrsp1n7v2Tfa9i6NVG2nBGFqw0vE9Gs0CCud6DNOPUnyTvoYYKuFwS1US8ll6wwyhPRe8F0pqWesHlSiITLcUvQtN1KFP8Osw1Bk1fR5nl8"
# #example_insight_id = "INSIGHT_ID"
#
# api_client = ApiClient(header_name="Authorization", header_value=access_token)
# insights_api = InsightsApi(api_client=api_client)
#
# insights_result = insights_api.get_insights()


from ceic_api_client.api_client import ApiClient
from ceic_api_client.apis.dictionary_api import DictionaryApi
from ceic_api_client.apis.series_api import SeriesApi
from ceic_api_client.apis.insights_api import InsightsApi


access_token = "oC6OX7cRKs386i8KhqkHVi5BOgRCVMAkUVmL1EQlrKz8qNztEkUoLhior6c2RYBmhkTUenYrLBM8Kzl4R7cu2c4ywikS1F4cevkHegEKEdIEhoyYMzkgKXAWXqwzHyC8"
example_insight_id = "INSIGHT_ID"

api_client = ApiClient(header_name="Authorization", header_value=access_token)
insights_api = InsightsApi(api_client=api_client)

insights_result = insights_api.get_insights().to_dict()
#print insights_result







#print (insights_result)
xml_data = dicttoxml.dicttoxml( insights_result, attr_type=False, root=True )
file_path = "C:\\Users\\rchethana\\Documents\\SDK_Automation\\XML-ActualData\\Metadata\insight_by_id.xml"
f = open( file_path, 'w' )
f.write( xml_data )
f.close()

tree = et.parse( file_path )
root = tree.getroot()
insight_id =[]

for name in root.findall( ".//item/id" ):
    insight_id.append( name.text )

sheet.cell( row=3, column=2 ).value = access_token
j = 3
for i in range(len(insight_id)):
    sheet.cell( row=j, column=3 ).value = insight_id[i]
    #print insight_id[i]
    j+=1
#print (len(insight_id))
#print (j)


for i in range(len(insight_id)):print(insight_id[i])
print len(insight_id)
